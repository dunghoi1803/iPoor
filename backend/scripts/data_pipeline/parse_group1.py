import csv
import json
import re
import sys
from pathlib import Path
from typing import Any
from unicodedata import normalize

import openpyxl


YEAR_PATTERN = re.compile(r"(19|20)\d{2}")


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = normalize("NFC", str(value))
    return " ".join(text.split()).strip()


def to_year(value: Any) -> int | None:
    if isinstance(value, (int, float)) and 1900 <= value <= 2100:
        return int(value)
    text = normalize_text(value)
    match = YEAR_PATTERN.search(text)
    if match:
        return int(match.group(0))
    return None


def read_config(config_path: Path) -> dict[str, Any]:
    return json.loads(config_path.read_text(encoding="utf-8"))


def find_label(row: list[Any], label_cols: list[int]) -> str:
    for col in label_cols:
        idx = col - 1
        if 0 <= idx < len(row):
            label = normalize_text(row[idx])
            if label:
                return label
    return ""


def row_has_numeric(row: list[Any], columns: list[int]) -> bool:
    for col in columns:
        idx = col - 1
        if 0 <= idx < len(row):
            value = row[idx]
            if isinstance(value, (int, float)):
                return True
    return False


def extract_year_columns(
    year_row: list[Any],
    start_col: int | None,
    end_col: int | None,
) -> list[tuple[int, int]]:
    pairs: list[tuple[int, int]] = []
    min_col = start_col or 1
    max_col = end_col or len(year_row)
    for col in range(min_col, max_col + 1):
        idx = col - 1
        if idx < 0 or idx >= len(year_row):
            continue
        year = to_year(year_row[idx])
        if year:
            pairs.append((col, year))
    return pairs


def normalize_key(value: str) -> str:
    return normalize_text(value).lower()


def resolve_metric_label(label: str, metric_aliases: dict[str, str]) -> str:
    if not label:
        return ""
    normalized = normalize_key(label)
    return metric_aliases.get(normalized, label)


def build_year_metric_map(
    year_row: list[Any],
    metric_row: list[Any],
    metric_fallback_row: list[Any] | None,
) -> dict[int, tuple[int, str]]:
    mapping: dict[int, tuple[int, str]] = {}
    current_year: int | None = None
    max_len = max(len(year_row), len(metric_row), len(metric_fallback_row or []))
    for col in range(1, max_len + 1):
        year = to_year(year_row[col - 1] if col - 1 < len(year_row) else None)
        if year:
            current_year = year
        metric_label = normalize_text(metric_row[col - 1] if col - 1 < len(metric_row) else None)
        if not metric_label and metric_fallback_row:
            metric_label = normalize_text(
                metric_fallback_row[col - 1] if col - 1 < len(metric_fallback_row) else None
            )
        if current_year and metric_label:
            mapping[col] = (current_year, metric_label)
    return mapping


def parse_sheet(
    workbook: openpyxl.Workbook,
    sheet_cfg: dict[str, Any],
    section_headers: set[str],
) -> list[dict[str, Any]]:
    name = sheet_cfg["name"]
    ws = workbook[name]
    layout = sheet_cfg.get("layout", "year_series")
    year_row_idx = sheet_cfg.get("year_row")
    data_start = sheet_cfg["data_start_row"]
    label_cols = sheet_cfg["label_cols"]
    blocks = sheet_cfg.get("blocks")
    metric = sheet_cfg.get("metric", "value")
    parent_label_col = sheet_cfg.get("parent_label_col")
    metric_row_idx = sheet_cfg.get("metric_row")
    metric_row_fallback_idx = sheet_cfg.get("metric_row_fallback")
    fixed_year = sheet_cfg.get("fixed_year")
    year_col_idx = sheet_cfg.get("year_col")
    metric_aliases = {
        normalize_key(key): value for key, value in sheet_cfg.get("metric_aliases", {}).items()
    }

    year_row = [cell.value for cell in ws[year_row_idx]] if year_row_idx else []
    metric_row = [cell.value for cell in ws[metric_row_idx]] if metric_row_idx else []
    metric_row_fallback = (
        [cell.value for cell in ws[metric_row_fallback_idx]]
        if metric_row_fallback_idx
        else None
    )
    if blocks:
        year_columns_by_block = []
        for block in blocks:
            cols = extract_year_columns(year_row, block["start_col"], block["end_col"])
            year_columns_by_block.append((block["metric"], cols))
    else:
        year_columns_by_block = [(metric, extract_year_columns(year_row, None, None))]

    indicator_title = ""
    for cell in ws[1]:
        indicator_title = normalize_text(cell.value)
        if indicator_title:
            break
    geo_version = "new_34" if name.endswith(".M") else "old_63"
    current_section = ""
    current_parent = ""
    last_label = ""
    empty_streak = 0
    rows_out: list[dict[str, Any]] = []

    all_year_columns = [col for _, cols in year_columns_by_block for col, _ in cols]
    year_metric_map = (
        build_year_metric_map(year_row, metric_row, metric_row_fallback) if metric_row else {}
    )

    for row_idx, row in enumerate(
        ws.iter_rows(min_row=data_start, values_only=True),
        start=data_start,
    ):
        row_values = list(row)
        if not any(cell is not None for cell in row_values):
            empty_streak += 1
            if empty_streak >= 3:
                break
            continue
        empty_streak = 0

        label = find_label(row_values, label_cols)
        if not label and layout == "row_year_metrics" and last_label:
            label = last_label
        if layout == "fixed_year_single":
            has_numeric = row_has_numeric(row_values, [sheet_cfg["value_col"]])
        elif layout == "fixed_year_metrics":
            metric_columns = [col for col in range(1, len(metric_row) + 1) if col not in label_cols]
            has_numeric = row_has_numeric(row_values, metric_columns)
        elif layout == "row_year_metrics":
            metric_columns = [
                col
                for col in range(1, len(metric_row) + 1)
                if col not in label_cols and col != year_col_idx
            ]
            has_numeric = row_has_numeric(row_values, metric_columns)
        elif layout == "year_metric_matrix":
            has_numeric = row_has_numeric(row_values, list(year_metric_map.keys()))
        else:
            has_numeric = row_has_numeric(row_values, all_year_columns)

        if not has_numeric and label:
            label_norm = normalize_text(label).lower()
            if label_norm in section_headers:
                current_section = label
                current_parent = ""
            elif parent_label_col:
                parent_label = normalize_text(row_values[parent_label_col - 1])
                if parent_label:
                    current_parent = parent_label
            continue

        if not label:
            continue
        last_label = label

        if layout == "fixed_year_metrics":
            for col in range(1, len(metric_row) + 1):
                if col in label_cols:
                    continue
                metric_label = normalize_text(metric_row[col - 1])
                if not metric_label:
                    continue
                metric_name = resolve_metric_label(metric_label, metric_aliases)
                value_idx = col - 1
                if value_idx < 0 or value_idx >= len(row_values):
                    continue
                value = row_values[value_idx]
                if value is None or not isinstance(value, (int, float)):
                    continue
                rows_out.append(
                    {
                        "sheet": name,
                        "indicator_code": name,
                        "indicator_title": indicator_title,
                        "geo_name": label,
                        "geo_parent": current_parent or "",
                        "section": current_section,
                        "metric": metric_name,
                        "year": fixed_year,
                        "value": value,
                        "geo_version": geo_version,
                    }
                )
            continue

        if layout == "fixed_year_single":
            value_idx = sheet_cfg["value_col"] - 1
            if 0 <= value_idx < len(row_values):
                value = row_values[value_idx]
                if value is not None and isinstance(value, (int, float)):
                    rows_out.append(
                        {
                            "sheet": name,
                            "indicator_code": name,
                            "indicator_title": indicator_title,
                            "geo_name": label,
                            "geo_parent": current_parent or "",
                            "section": current_section,
                            "metric": metric,
                            "year": fixed_year,
                            "value": value,
                            "geo_version": geo_version,
                        }
                    )
            continue

        if layout == "row_year_metrics":
            if not metric_row:
                continue
            year_value = None
            if year_col_idx:
                if 0 <= year_col_idx - 1 < len(row_values):
                    year_value = to_year(row_values[year_col_idx - 1])
            if not year_value:
                continue
            for col in range(1, len(metric_row) + 1):
                if col in label_cols or col == year_col_idx:
                    continue
                metric_label = normalize_text(metric_row[col - 1])
                if not metric_label:
                    continue
                metric_name = resolve_metric_label(metric_label, metric_aliases)
                value_idx = col - 1
                if value_idx < 0 or value_idx >= len(row_values):
                    continue
                value = row_values[value_idx]
                if value is None or not isinstance(value, (int, float)):
                    continue
                rows_out.append(
                    {
                        "sheet": name,
                        "indicator_code": name,
                        "indicator_title": indicator_title,
                        "geo_name": label,
                        "geo_parent": current_parent or "",
                        "section": current_section,
                        "metric": metric_name,
                        "year": year_value,
                        "value": value,
                        "geo_version": geo_version,
                    }
                )
            continue

        if layout == "year_metric_matrix":
            for col, (year, metric_label) in year_metric_map.items():
                value_idx = col - 1
                if value_idx < 0 or value_idx >= len(row_values):
                    continue
                value = row_values[value_idx]
                if value is None or not isinstance(value, (int, float)):
                    continue
                metric_name = resolve_metric_label(metric_label, metric_aliases)
                rows_out.append(
                    {
                        "sheet": name,
                        "indicator_code": name,
                        "indicator_title": indicator_title,
                        "geo_name": label,
                        "geo_parent": current_parent or "",
                        "section": current_section,
                        "metric": metric_name,
                        "year": year,
                        "value": value,
                        "geo_version": geo_version,
                    }
                )
            continue

        for block_metric, year_cols in year_columns_by_block:
            for col, year in year_cols:
                value_idx = col - 1
                if value_idx < 0 or value_idx >= len(row_values):
                    continue
                value = row_values[value_idx]
                if value is None:
                    continue
                if not isinstance(value, (int, float)):
                    continue
                rows_out.append(
                    {
                        "sheet": name,
                        "indicator_code": name,
                        "indicator_title": indicator_title,
                        "geo_name": label,
                        "geo_parent": current_parent or "",
                        "section": current_section,
                        "metric": block_metric,
                        "year": year,
                        "value": value,
                        "geo_version": geo_version,
                    }
                )

    return rows_out


def main() -> None:
    root = Path(__file__).resolve().parents[2]
    config_path = root / "scripts" / "data_pipeline" / "group1_config.json"
    if len(sys.argv) > 1:
        config_path = Path(sys.argv[1])
    config = read_config(config_path)
    workbook_path = root.parent / "FE" / "data" / "So lieu ve ban do 27 Nov 2025_2.xlsx"
    output_path = root.parent / config["output"]
    output_path.parent.mkdir(parents=True, exist_ok=True)

    section_headers = {normalize_text(value).lower() for value in config["section_headers"]}
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)

    rows_out: list[dict[str, Any]] = []
    for sheet_cfg in config["sheets"]:
        rows_out.extend(parse_sheet(wb, sheet_cfg, section_headers))

    fieldnames = [
        "sheet",
        "indicator_code",
        "indicator_title",
        "geo_name",
        "geo_parent",
        "section",
        "metric",
        "year",
        "value",
        "geo_version",
    ]

    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows_out)

    print(f"Wrote {len(rows_out)} rows to {output_path}")


if __name__ == "__main__":
    main()
