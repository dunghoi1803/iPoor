from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from .. import deps, models, schemas
from ..config import get_settings
from ..constants import CollectionStatus, DEFAULT_PAGE_LIMIT, MAX_PAGE_LIMIT, PovertyStatus
from ..utils.file_naming import (
    FILENAME_MAX_LENGTH,
    FILENAME_SEPARATOR,
    build_household_prefix,
)
from ..utils.household_code import generate_household_code
from ..utils.text import normalize_header, normalize_text

router = APIRouter(prefix="/data-collections", tags=["data_collections"])
MAX_UPLOAD_BYTES = 50 * 1024 * 1024
ALLOWED_EXTENSIONS = {".xlsx", ".csv"}
PDF_SUBDIR = "data-collections"
UPLOAD_CHUNK_SIZE = 1024 * 1024
HEADER_ROW_INDEX = 2
SUBHEADER_ROW_INDEX = 3
DATA_ROW_INDEX = 6
ALLOWED_PDF_EXTENSIONS = {".pdf"}
HEADER_ALIASES = {
    "variables": "variables",
    "name": "name",
    "dob": "dob",
    "gender": "gender",
    "ethnic": "ethnic",
    "address_line": "address_line",
    "id_num": "ID_num",
    "family_mem": "family_mem",
    "classified_before_check": "classified_before_check",
    "b1_score": "B1_score",
    "b2_score": "B2_score",
    "income_per_capita": "income_per_capita",
    "classified_after_check": "classified_after_check",
    "description": "description",
    "note": "note",
    "pdf_url": "pdf_url",
    "province": "province",
    "district": "district",
    "commune": "commune",
    "village": "village",
    "date_check": "date_check",
    "official_check": "official_check",
    "khu vực": "area",
    "tài liệu đính kèm": "pdf_url",
    "ngày rà soát": "date_check",
    "cán bộ rà soát": "official_check",
}
POVERTY_STATUS_MAP = {
    "1": PovertyStatus.POOR,
    "2": PovertyStatus.NEAR_POOR,
    "3": PovertyStatus.ESCAPED,
    "4": PovertyStatus.AT_RISK,
    "hộ nghèo": PovertyStatus.POOR,
    "hộ cận nghèo": PovertyStatus.NEAR_POOR,
    "hộ thoát nghèo": PovertyStatus.ESCAPED,
    "hộ có khả năng tái nghèo": PovertyStatus.AT_RISK,
    "nghèo": PovertyStatus.POOR,
    "cận nghèo": PovertyStatus.NEAR_POOR,
    "thoát nghèo": PovertyStatus.ESCAPED,
    "tái nghèo": PovertyStatus.AT_RISK,
    "nguy cơ tái nghèo": PovertyStatus.AT_RISK,
    "có khả năng tái nghèo": PovertyStatus.AT_RISK,
}
REQUIRED_FIELDS = {
    "name": "Họ và tên",
    "ID_num": "CCCD",
    "family_mem": "Số thành viên",
    "classified_after_check": "Kết luận rà soát",
    "B1_score": "Điểm B1",
    "B2_score": "Điểm B2",
    "commune": "Xã",
    "village": "Thôn",
    "date_check": "Ngày rà soát",
    "official_check": "Cán bộ rà soát",
}
ATTACHMENT_LABEL = "Tài liệu đính kèm"


def parse_date(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d")
    except ValueError:
        pass
    try:
        return datetime.strptime(str(value).strip(), "%d/%m/%Y")
    except ValueError:
        return None


def parse_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def parse_poverty_status(value: Any) -> str | None:
    text = normalize_text(value).casefold()
    if not text:
        return None
    return POVERTY_STATUS_MAP.get(text)


def build_error(row_number: int, column: str, message: str) -> schemas.DataCollectionUploadError:
    return schemas.DataCollectionUploadError(row=row_number, column=column, message=message)


def get_cell_value(row: list[Any], column_index: int) -> Any:
    idx = column_index - 1
    return row[idx] if 0 <= idx < len(row) else None


def build_column_map(
    header_row: list[Any],
    subheader_row: list[Any],
) -> dict[str, int]:
    column_map: dict[str, int] = {}
    max_len = max(len(header_row), len(subheader_row))
    for index in range(max_len):
        header = header_row[index] if index < len(header_row) else None
        subheader = subheader_row[index] if index < len(subheader_row) else None
        column_index = index + 1
        normalized_header = normalize_header(header)
        key = HEADER_ALIASES.get(normalized_header)
        if not key:
            normalized_subheader = normalize_header(subheader)
            key = HEADER_ALIASES.get(normalized_subheader)
        if key and key not in column_map:
            column_map[key] = column_index
    return column_map


def parse_excel_rows(file_bytes: bytes) -> tuple[list[Any], list[Any], list[list[Any]]]:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    header_row = [cell.value for cell in sheet[HEADER_ROW_INDEX]]
    subheader_row = [cell.value for cell in sheet[SUBHEADER_ROW_INDEX]]
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(min_row=DATA_ROW_INDEX, values_only=True):
        rows.append(list(row))
    return header_row, subheader_row, rows


def parse_csv_rows(file_bytes: bytes) -> tuple[list[Any], list[Any], list[list[Any]]]:
    import csv

    text = file_bytes.decode("utf-8-sig", errors="ignore")
    rows: list[list[Any]] = []
    for row in csv.reader(text.splitlines()):
        rows.append([cell.strip() for cell in row])
    header_row = rows[HEADER_ROW_INDEX - 1] if len(rows) >= HEADER_ROW_INDEX else []
    subheader_row = rows[SUBHEADER_ROW_INDEX - 1] if len(rows) >= SUBHEADER_ROW_INDEX else []
    data_rows = rows[DATA_ROW_INDEX - 1 :] if len(rows) >= DATA_ROW_INDEX else []
    return header_row, subheader_row, data_rows


def normalize_filename_key(value: str) -> str:
    normalized = normalize("NFC", value).strip()
    return normalized.casefold()


def ensure_unique_path(target_dir: Path, filename: str) -> Path:
    target_dir.mkdir(parents=True, exist_ok=True)
    file_path = target_dir / filename
    if not file_path.exists():
        return file_path
    extension = file_path.suffix
    stem = file_path.stem
    counter = 1
    while True:
        candidate = f"{stem}{FILENAME_SEPARATOR}{counter}{extension}"
        candidate = candidate[:FILENAME_MAX_LENGTH]
        file_path = target_dir / candidate
        if not file_path.exists():
            return file_path
        counter += 1


def save_upload_file(upload_file: UploadFile, dest_path: Path) -> None:
    upload_file.file.seek(0)
    with dest_path.open("wb") as output:
        while True:
            chunk = upload_file.file.read(UPLOAD_CHUNK_SIZE)
            if not chunk:
                break
            output.write(chunk)


@router.post("/upload", response_model=schemas.DataCollectionUploadResult)
def upload_data_collection(
    file: UploadFile = File(...),  # noqa: B008
    pdf_files: list[UploadFile] | None = File(None),
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_user),
) -> schemas.DataCollectionUploadResult:
    if not file.filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Missing file")
    extension = "." + file.filename.split(".")[-1].lower()
    if extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unsupported file type")
    file_bytes = file.file.read()
    if len(file_bytes) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File too large")

    if extension == ".xlsx":
        header_row, subheader_row, rows = parse_excel_rows(file_bytes)
    else:
        header_row, subheader_row, rows = parse_csv_rows(file_bytes)

    column_map = build_column_map(header_row, subheader_row)
    missing_columns = [label for key, label in REQUIRED_FIELDS.items() if key not in column_map]
    if missing_columns:
        joined = ", ".join(missing_columns)
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Missing columns: {joined}",
        )

    pdf_lookup: dict[str, UploadFile] = {}
    if pdf_files:
        for pdf_file in pdf_files:
            if not pdf_file.filename:
                continue
            suffix = Path(pdf_file.filename).suffix.lower()
            if suffix not in ALLOWED_PDF_EXTENSIONS:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail="Only PDF files are allowed",
                )
            key = normalize_filename_key(Path(pdf_file.filename).name)
            pdf_lookup[key] = pdf_file
    errors: list[schemas.DataCollectionUploadError] = []
    row_summaries: list[schemas.DataCollectionUploadRow] = []
    valid_count = 0
    error_count = 0
    for index, row in enumerate(rows, start=DATA_ROW_INDEX):
        values = {
            key: get_cell_value(row, col) for key, col in column_map.items()
        }
        row_errors: list[schemas.DataCollectionUploadError] = []
        for key, label in REQUIRED_FIELDS.items():
            if normalize_text(values.get(key)) == "":
                row_errors.append(build_error(index, label, "Thiếu dữ liệu"))

        if all(normalize_text(value) == "" for value in values.values()):
            continue

        members = parse_int(values.get("family_mem"))
        if members is None:
            row_errors.append(build_error(index, REQUIRED_FIELDS["family_mem"], "Không hợp lệ"))

        b1_score = parse_int(values.get("B1_score"))
        if b1_score is None:
            row_errors.append(build_error(index, REQUIRED_FIELDS["B1_score"], "Không hợp lệ"))

        b2_score = parse_int(values.get("B2_score"))
        if b2_score is None:
            row_errors.append(build_error(index, REQUIRED_FIELDS["B2_score"], "Không hợp lệ"))

        poverty_status = parse_poverty_status(values.get("classified_after_check"))
        if poverty_status is None:
            row_errors.append(build_error(index, REQUIRED_FIELDS["classified_after_check"], "Không hợp lệ"))

        collected_at = parse_date(values.get("date_check"))
        if values.get("date_check") and not collected_at:
            row_errors.append(build_error(index, REQUIRED_FIELDS["date_check"], "Sai định dạng ngày"))

        birth_date = parse_date(values.get("dob"))
        id_card = normalize_text(values.get("ID_num"))
        if id_card:
            exists = (
                db.query(models.Household)
                .filter(models.Household.id_card == id_card)
                .first()
            )
            if exists:
                row_errors.append(build_error(index, REQUIRED_FIELDS["ID_num"], "CCCD đã tồn tại"))
        pdf_ref = normalize_text(values.get("pdf_url"))
        if pdf_ref:
            pdf_key = normalize_filename_key(Path(pdf_ref).name)
            pdf_file = pdf_lookup.get(pdf_key)
            if not pdf_file:
                row_errors.append(build_error(index, ATTACHMENT_LABEL, "Không tìm thấy file PDF"))

        row_summary = schemas.DataCollectionUploadRow(
            row=index,
            name=normalize_text(values.get("name")) or None,
            id_card=id_card or None,
            poverty_status=normalize_text(values.get("classified_after_check")) or None,
            birth_date=normalize_text(values.get("dob")) or None,
            gender=normalize_text(values.get("gender")) or None,
            ethnic=normalize_text(values.get("ethnic")) or None,
            address_line=normalize_text(values.get("address_line")) or None,
            family_mem=members,
            classified_before_check=normalize_text(values.get("classified_before_check")) or None,
            b1_score=b1_score,
            b2_score=b2_score,
            income_per_capita=parse_float(values.get("income_per_capita")),
            area=normalize_text(values.get("area")) or None,
            description=normalize_text(values.get("description")) or None,
            note=normalize_text(values.get("note")) or None,
            pdf_url=pdf_ref or None,
            province=normalize_text(values.get("province")) or None,
            district=normalize_text(values.get("district")) or None,
            commune=normalize_text(values.get("commune")) or None,
            village=normalize_text(values.get("village")) or None,
            date_check=normalize_text(values.get("date_check")) or None,
            official_check=normalize_text(values.get("official_check")) or None,
            valid=len(row_errors) == 0,
            errors=row_errors,
        )
        row_summaries.append(row_summary)

        if row_errors:
            errors.extend(row_errors)
            error_count += 1
            continue
        valid_count += 1

    return schemas.DataCollectionUploadResult(
        validRecords=valid_count,
        errorRecords=error_count,
        errors=errors,
        rows=row_summaries,
    )


@router.post("/commit", response_model=schemas.HouseholdRead, status_code=status.HTTP_201_CREATED)
def commit_data_collection_row(
    name: str = Form(...),
    dob: str | None = Form(None),
    gender: str | None = Form(None),
    ethnic: str | None = Form(None),
    address_line: str | None = Form(None),
    id_num: str = Form(...),
    family_mem: int = Form(...),
    classified_before_check: str | None = Form(None),
    b1_score: int = Form(...),
    b2_score: int = Form(...),
    income_per_capita: float | None = Form(None),
    classified_after_check: str = Form(...),
    area: str | None = Form(None),
    description: str | None = Form(None),
    note: str | None = Form(None),
    pdf_url: str | None = Form(None),
    province: str = Form(...),
    district: str = Form(...),
    commune: str = Form(...),
    village: str = Form(...),
    date_check: str = Form(...),
    official_check: str = Form(...),
    pdf_file: UploadFile | None = File(None),  # noqa: B008
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_user),
) -> models.Household:
    existing = db.query(models.Household).filter(models.Household.id_card == id_num).first()
    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail="Household already exists",
        )
    poverty_status = parse_poverty_status(classified_after_check)
    if not poverty_status:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid poverty status",
        )
    collected_at = parse_date(date_check)
    if date_check and not collected_at:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Invalid date format",
        )
    birth_date = parse_date(dob)

    attachment_url = None
    if pdf_url:
        if not pdf_file:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Missing PDF file",
            )
        if Path(pdf_file.filename or "").suffix.lower() not in ALLOWED_PDF_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only PDF files are allowed",
            )
        settings = get_settings()
        target_dir = Path(settings.upload_dir) / PDF_SUBDIR
        household_code = generate_household_code(db)
        prefix = build_household_prefix(household_code, poverty_status, name, id_num)
        extension = Path(pdf_file.filename or "").suffix or ".pdf"
        raw_name = f"{prefix}{extension}"
        safe_name = raw_name[:FILENAME_MAX_LENGTH]
        file_path = ensure_unique_path(target_dir, safe_name)
        save_upload_file(pdf_file, file_path)
        attachment_url = f"/files/{PDF_SUBDIR}/{quote(file_path.name)}"
    else:
        household_code = generate_household_code(db)

    household = models.Household(
        household_code=household_code,
        head_name=normalize_text(name),
        birth_date=birth_date.date() if birth_date else None,
        gender=normalize_text(gender) or None,
        ethnicity=normalize_text(ethnic) or None,
        id_card=normalize_text(id_num) or None,
        members_count=family_mem,
        income_per_capita=income_per_capita,
        poverty_status=poverty_status,
        score_b1=b1_score,
        score_b2=b2_score,
        note=normalize_text(description) or None,
        remark=normalize_text(note) or None,
        area=normalize_text(area) or None,
        village=normalize_text(village) or None,
        officer=normalize_text(official_check) or None,
        commune=normalize_text(commune) or "Chưa rõ",
        province=normalize_text(province) or "Chưa rõ",
        district=normalize_text(district) or "Chưa rõ",
        address_line=normalize_text(address_line) or None,
        attachment_url=attachment_url,
        last_surveyed_at=collected_at.date() if collected_at else None,
    )
    db.add(household)
    db.flush()

    notes = f"Before: {normalize_text(classified_before_check)}; After: {normalize_text(classified_after_check)}"
    collection = models.DataCollection(
        household_id=household.id,
        collector_id=current_user.id,
        status=CollectionStatus.SUBMITTED,
        notes=notes,
        collected_at=collected_at,
    )
    db.add(collection)
    db.commit()
    db.refresh(household)
    return household


@router.get("", response_model=list[schemas.DataCollectionRead])
def list_data_collections(
    household_id: int | None = None,
    status_filter: CollectionStatus | None = None,
    skip: int = 0,
    limit: int = Query(DEFAULT_PAGE_LIMIT, le=MAX_PAGE_LIMIT),
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_user),
) -> list[models.DataCollection]:
    query = db.query(models.DataCollection)
    if household_id:
        query = query.filter(models.DataCollection.household_id == household_id)
    if status_filter:
        query = query.filter(models.DataCollection.status == status_filter)
    return query.order_by(models.DataCollection.created_at.desc()).offset(skip).limit(limit).all()


@router.post("", response_model=schemas.DataCollectionRead, status_code=status.HTTP_201_CREATED)
def create_data_collection(
    payload: schemas.DataCollectionCreate,
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_user),
) -> models.DataCollection:
    household = (
        db.query(models.Household).filter(models.Household.id == payload.household_id).first()
    )
    if not household:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Household not found")
    collection = models.DataCollection(**payload.model_dump())
    db.add(collection)
    db.commit()
    db.refresh(collection)
    return collection


@router.put("/{collection_id}", response_model=schemas.DataCollectionRead)
def update_data_collection(
    collection_id: int,
    payload: schemas.DataCollectionUpdate,
    db: Session = Depends(deps.get_db),
    current_user: models.User = Depends(deps.get_current_user),
) -> models.DataCollection:
    collection = (
        db.query(models.DataCollection)
        .filter(models.DataCollection.id == collection_id)
        .first()
    )
    if not collection:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Data collection not found")
    for key, value in payload.model_dump(exclude_unset=True).items():
        setattr(collection, key, value)
    db.commit()
    db.refresh(collection)
    return collection
